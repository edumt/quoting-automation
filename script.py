#import os
#import datetime
#from openpyxl import Workbook
from openpyxl import load_workbook
from docx import Document

CONST_SHEET_VERSION = '13-JULHO-2021'
CONST_CONSUMPTION_CELL = 'D18'
CONST_NAME_CELL = 'C4'
CONST_PERCENTAGE_INCREASE_CELL = 'H44'
CONST_OPTION_1_PANEL_NUMBER_CELL = 'D11'
CONST_OPTION_2_PANEL_NUMBER_CELL = 'F11'
CONST_OPTION_3_PANEL_NUMBER_CELL = 'H11'
CONST_BUDGET_CELL = 'B3:H32' 

statesDict = {
    "ES": "ESPÍRITO SANTO",
    "RJ": "RIO DE JANEIRO",
    "BA": "BAHIA",
    "MG": "MINAS GERAIS"
}

class Client:
    def __init__(self, name, consumption, state, city = ''):
        self.name = name.upper()
        self.consumption = consumption # Average monthly energy consumption [kWh/month]
        self.state = state.upper() # State initials, e.g. ES
        self.city = city.upper() # (District-)City, e.g. PRAIA DE ITAPARICA-VILA VELHA
    
    def getFormulaSheet(self):
        self.sheet = load_workbook('../excel_template.xlsx')
        
    def getDataSheet(self):
        self.sheet = load_workbook('../excel_template.xlsx', data_only=True)
    
    def populateSheet(self):
        #print(sheet.sheetnames)
        self.sheet['HCONSUMO'][CONST_CONSUMPTION_CELL].value = self.consumption
        self.sheet['HCONSUMO'][CONST_NAME_CELL].value = ('CLIENTE: {name}').format(name=self.name)
        self.setPanelsQuantity()
        
    def setPanelsQuantity(self):
        # TO DO: dinamically change number of PV panels
        self.sheet['Preço SFCR-GROWATT PHONO 450Wp'][CONST_OPTION_2_PANEL_NUMBER_CELL].value =\
        (self.sheet['Preço SFCR-GROWATT PHONO 450Wp'][CONST_OPTION_1_PANEL_NUMBER_CELL].value + '+1')
            
        self.sheet['Preço SFCR-GROWATT PHONO 450Wp'][CONST_OPTION_3_PANEL_NUMBER_CELL].value =\
        (self.sheet['Preço SFCR-GROWATT PHONO 450Wp'][CONST_OPTION_2_PANEL_NUMBER_CELL].value + '+1')

    def saveSheet(self):
        # TO DO: learn mkdir() and/or save() best practices, what to do when dir/file aready exists, exception handling etc
        #os.mkdir(('../{name}').format(name=client.name))
        #self.sheet.save(('../{name}/GERADORES-{name}-ALDO-{version}.xlsx').format(name=self.name, version=CONST_SHEET_VERSION))
        self.sheet.save('../test.xlsx')
        #pass
    
    def generateSheet(self):
        self.getFormulaSheet()
        self.populateSheet()
        self.saveSheet()
        print('Success! Sheet generated.')

    def copyBudgetArea(self):
        self.getDataSheet()

    def generateReport(self):
        pass

# TO DO: CLI and, eventually, a GUI
test_client = Client('Eduardo Moura Tavares', 5000, 'ES', 'Praia de Itaparica-Vila Velha')
#print(vars(test_client))
test_client.generateSheet()